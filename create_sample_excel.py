"""
create_sample_excel.py
점포 데이터 샘플 엑셀 파일 생성 스크립트.
최초 1회만 실행하면 됩니다.
이후에는 store_profiles.xlsx 파일을 직접 수정해서 사용하세요.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "점포데이터"

headers = ["점포명", "연간골프매출(억)", "구매고객수(명)", "평균객단가(만원)", "우수고객구성비(%)", "평균연령(세)", "점포특성"]
data = [
    ["잠실점",   350, 20000, 80, 45, 56, "롯데월드몰 이용가능"],
    ["본점",     250, 15000, 90, 50, 47, "외국인 고객 많음"],
    ["부산본점", 250, 18000, 80, 40, 51, "바닷가"],
    ["인천점",   200, 12000, 80, 35, 43, "상권단독"],
    ["동탄점",   150, 10000, 80, 40, 39, "젊은점포"],
    ["노원점",   130,  8000, 60, 32, 54, "포켓상권"],
    ["영등포점", 100,  7000, 40, 30, 60, "경쟁열위"],
    ["광복점",    80,  6000, 40, 28, 55, "점포는큼"],
    ["미아점",    40,  4000, 30, 21, 65, "행사효율 좋음"],
    ["건대점",    40,  4000, 20, 15, 43, "골프 부진"],
]

# 헤더 스타일
header_fill = PatternFill("solid", fgColor="064E3B")
header_font = Font(bold=True, color="FFFFFF", size=11)
thin = Side(style="thin", color="D1FAE5")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = border

# 데이터
alt_fill = PatternFill("solid", fgColor="F0FDF4")
for row_idx, row in enumerate(data, 2):
    for col_idx, val in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.alignment = Alignment(horizontal="center" if col_idx != len(headers) else "left", vertical="center")
        cell.border = border
        if row_idx % 2 == 0:
            cell.fill = alt_fill

# 컬럼 너비
col_widths = [12, 18, 16, 18, 18, 14, 20]
for i, w in enumerate(col_widths, 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
ws.row_dimensions[1].height = 24

wb.save("store_profiles.xlsx")
print("store_profiles.xlsx 생성 완료!")
